# -*- coding: utf-8 -*-
"""
build_data.py — xlsx -> data.json, avec assertions.

Le pipeline doit reproduire EXACTEMENT les chiffres de reference du plan
(section 2). Toute derive doit faire echouer le script bruyamment.

Lecture volontairement faite avec openpyxl et des index de ligne explicites :
`Credit bancaire.xlsx` empile deux tables dans la meme feuille, et laisser
pandas deviner l'en-tete produit un bloc B decale d'une ligne.

Usage :  py -3 pipeline/build_data.py
"""
import csv
import io
import json
import os
import sys
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "data", "raw")
OUT = os.path.join(ROOT, "data", "data.json")
EMBED = os.path.join(ROOT, "js", "data.js")
MAPPING = os.path.join(ROOT, "data", "mapping_secteurs.csv")

ANNEES_REF = [2015, 2016, 2017]
ANNEE_DEFAUT = 2017

# --------------------------------------------------------------------------
# outils
# --------------------------------------------------------------------------

def d2(x, n=2):
    """Arrondi 'commercial' (half-up), identique a ce qu'affichera toFixed()."""
    return float(Decimal(repr(float(x))).quantize(Decimal("1." + "0" * n), rounding=ROUND_HALF_UP))


def arrondir_a_somme_nulle(vals):
    """Arrondit a l'entier en garantissant que la somme affichee reste nulle.

    Arrondir chaque ecart independamment peut donner une somme de -1 ou +1
    (cas 2017 hors extraction) : le waterfall ne boucle plus et le graphique ment.
    On corrige d'une unite les valeurs dont l'arrondi coute le moins.
    """
    r = [int(d2(v, 0)) for v in vals]
    s = sum(r)
    while s != 0:
        pas = -1 if s > 0 else 1
        i = min(range(len(vals)),
                key=lambda k: abs(vals[k] - (r[k] + pas)) - abs(vals[k] - r[k]))
        r[i] += pas
        s += pas
    return r


CHECKS = []

def check(label, got, want, tol=0.0):
    ok = abs(float(got) - float(want)) <= tol
    CHECKS.append((ok, label, got, want))
    if not ok:
        raise AssertionError(f"CONTROLE ECHOUE — {label}: obtenu {got!r}, attendu {want!r} (tol {tol})")


def check_eq(label, got, want):
    ok = got == want
    CHECKS.append((ok, label, got, want))
    if not ok:
        raise AssertionError(f"CONTROLE ECHOUE — {label}: obtenu {got!r}, attendu {want!r}")


# --------------------------------------------------------------------------
# 1. mapping
# --------------------------------------------------------------------------

def lire_mapping():
    postes, branches_extraction, non_couvert = {}, set(), []
    with open(MAPPING, encoding="utf-8") as f:
        for row in csv.DictReader(f, delimiter=";"):
            poste = row["poste_credit"].strip()
            branche = row["branche_va"].strip()
            if row["flag"].strip() == "extraction":
                branches_extraction.add(branche)
            if poste == "(non couvert)":
                non_couvert.append(branche)
            else:
                postes.setdefault(poste, []).append(branche)
    return postes, branches_extraction, non_couvert


# --------------------------------------------------------------------------
# 2. comptes nationaux
# --------------------------------------------------------------------------

def lire_comptes():
    ws = openpyxl.load_workbook(os.path.join(RAW, "Comptes Nationaux.xlsx"), data_only=True)["sheet"]
    entetes = [(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(entetes)}
    annees = {}
    for r in range(2, ws.max_row + 1):
        an = int(ws.cell(r, 1).value)
        annees[an] = {h: ws.cell(r, idx[h]).value for h in entetes if h}
    return annees, entetes


# --------------------------------------------------------------------------
# 3. credit bancaire (deux blocs empiles)
# --------------------------------------------------------------------------

MOIS = {"janv": 1, "févr": 2, "mars": 3, "avr": 4, "mai": 5, "juin": 6,
        "juil": 7, "août": 8, "sept": 9, "oct": 10, "nov": 11, "déc": 12}

def _mois(label):
    mot, an = label.strip().rsplit(" ", 1)
    return int(an), MOIS[mot.replace(".", "").strip().lower()]


def lire_credit():
    ws = openpyxl.load_workbook(os.path.join(RAW, "Crédit bancaire.xlsx"), data_only=True)["Sheet"]

    # -- bloc A : lignes 2..16, colonnes A/B/C. Le fichier est ANTICHRONOLOGIQUE.
    bloc_a = []
    for r in range(2, 17):
        lab = ws.cell(r, 1).value
        an, mo = _mois(lab)
        bloc_a.append({"label": lab.strip(), "iso": f"{an:04d}-{mo:02d}",
                       "ct": float(ws.cell(r, 2).value), "mlt": float(ws.cell(r, 3).value)})
    bloc_a.sort(key=lambda d: d["iso"])          # remise en ordre chronologique

    # -- bloc B : lignes 22..27 (secteurs + residu), ligne 28 = TOTAL (jamais dans les parts).
    # Les colonnes D/E/F sont une HYPOTHESE saisie a la main ("Hyp. % Court terme (saisie)"),
    # pas une donnee publiee : on les lit pour le controle, on ne les exporte pas.
    entete_d = str(ws.cell(21, 4).value or "")
    assert "saisie" in entete_d.lower(), (
        "La colonne D du bloc B ne porte plus la mention 'saisie' : verifier avant de l'exporter, "
        f"en-tete lu = {entete_d!r}")

    postes, residu = [], None
    for r in range(22, 28):
        nom = str(ws.cell(r, 1).value).strip()
        part = float(ws.cell(r, 2).value)
        enc = float(ws.cell(r, 3).value)
        if nom.lower().startswith("consommation"):
            residu = {"nom": "Consommation et autres", "part": part, "encours": enc}
        else:
            court = nom.split(" (")[0].strip()
            postes.append({"id": court.lower().replace("ê", "e").replace(" ", "-"),
                           "nom": court, "part_publiee": part, "encours": enc})

    assert str(ws.cell(28, 1).value).strip().upper() == "TOTAL", "ligne 28 attendue = TOTAL"
    total = float(ws.cell(28, 3).value)
    ct_total = float(ws.cell(28, 5).value)
    mlt_total = float(ws.cell(28, 6).value)
    return bloc_a, postes, residu, total, ct_total, mlt_total


# --------------------------------------------------------------------------
# 4. FINAN
# --------------------------------------------------------------------------

def lire_finan():
    ws = openpyxl.load_workbook(os.path.join(RAW, "FINAN.xlsx"), data_only=True)["SHEET"]
    # colonnes : 2-5 = nombre d'institutions, 6-8 = nombre d'agences, 9 = "Country wide" (non defini -> ignore)
    cols = {"banques_inst": 2, "coop_inst": 3, "imf_inst": 4, "autres_inst": 5,
            "banques_agences": 6, "coop_agences": 7, "imf_agences": 8}
    lignes = []
    for r in range(2, ws.max_row + 1):
        an = ws.cell(r, 1).value
        if an in (None, ""):
            continue
        rec = {"annee": int(an)}
        for k, c in cols.items():
            v = ws.cell(r, c).value
            rec[k] = None if v in (None, "") else int(v)   # cellules vides = chaines '', pas NaN
        lignes.append(rec)
    lignes.sort(key=lambda d: d["annee"])
    return lignes


# --------------------------------------------------------------------------
# 5. scenarios
# --------------------------------------------------------------------------

def construire_scenarios(comptes, postes_map, branches_extraction, postes_credit, total_credit, residu):
    ventile = sum(p["encours"] for p in postes_credit)
    scenarios = {}
    for annee in ANNEES_REF:
        va_an = comptes[annee]
        for extraction in ("incl", "excl"):
            va = {}
            for poste, branches in postes_map.items():
                gardees = [b for b in branches
                           if not (extraction == "excl" and b in branches_extraction)]
                va[poste] = sum(float(va_an[b]) for b in gardees)
            va_total = sum(va.values())

            for mode_residu in ("excl", "repart"):
                base = ventile if mode_residu == "excl" else total_credit
                secteurs, ecarts_exacts = [], []
                for p in postes_credit:
                    part_va = va[p["nom"]] / va_total
                    credit = p["encours"] if mode_residu == "excl" else \
                        p["encours"] + residu["encours"] * part_va
                    part_credit = credit / base
                    ecart = credit - part_va * base
                    ecarts_exacts.append(ecart)
                    nom = p["nom"]
                    if nom == "Industries" and extraction == "excl":
                        nom = "Industries (manuf.)"
                    secteurs.append({
                        "id": p["id"], "nom": nom,
                        "va": round(va[p["nom"]], 2),
                        "part_va": round(part_va, 8),
                        "credit": round(credit, 2),
                        "part_credit": round(part_credit, 8),
                        "ifr": round(part_credit / part_va, 8),
                        "ecart": round(ecart, 4),
                    })
                # etiquette entiere dont la somme reste nulle (geometrie = valeur exacte)
                for s, aff in zip(secteurs, arrondir_a_somme_nulle(ecarts_exacts)):
                    s["ecart_aff"] = aff
                secteurs.sort(key=lambda s: -s["ifr"])
                cle = f"{annee}|{extraction}|{mode_residu}"
                scenarios[cle] = {
                    "annee": annee, "extraction": extraction, "residu": mode_residu,
                    "va_total": round(va_total, 2), "base_credit": round(base, 2),
                    "secteurs": secteurs,
                }
                # la somme des ecarts vaut 0 par construction : sinon le waterfall ment.
                # Controle sur les valeurs exactes, puis sur les valeurs arrondies telles
                # qu'elles seront affichees (l'ecran 2 additionne ces dernieres).
                check(f"somme des ecarts nulle [{cle}]", sum(ecarts_exacts), 0, tol=1e-9)
                check(f"somme des ecarts affiches nulle [{cle}]",
                      sum(s["ecart_aff"] for s in secteurs), 0)
    return scenarios, ventile


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------

def main():
    postes_map, branches_extraction, non_couvert = lire_mapping()
    comptes, entetes = lire_comptes()
    bloc_a, postes_credit, residu, total_credit, ct_total, mlt_total = lire_credit()
    finan = lire_finan()

    # --- coherence des sources -------------------------------------------------
    va2017 = comptes[2017]
    pib_cf = float(va2017["PIB aux coût de facteurs"])
    check("PIB cout des facteurs 2017", pib_cf, 229624)
    check("PIB marche = PIB cf + impots nets 2017",
          pib_cf + float(va2017["Impôts nets de subvention sur produits"]),
          float(va2017["PIB aux prix du marche"]), tol=0.5)
    check("total credit publie", total_credit, 75146.23, tol=0.01)
    check("residu non ventile", residu["encours"], 30809.9543, tol=0.01)
    check("part du residu (%)", d2(residu["encours"] / total_credit * 100, 1), 41.0)
    check("bloc A : 15 mois", len(bloc_a), 15)
    check("bloc A ordre chronologique", bloc_a[0]["iso"] == "2020-01" and bloc_a[-1]["iso"] == "2021-03", True)

    scenarios, ventile = construire_scenarios(
        comptes, postes_map, branches_extraction, postes_credit, total_credit, residu)
    check("encours ventile", ventile, 44336.2757, tol=0.01)

    # --- section 2 : scenario A (2017, extraction incluse, residu exclu) --------
    A = {s["nom"]: s for s in scenarios["2017|incl|excl"]["secteurs"]}
    check("VA totale des 5 secteurs", scenarios["2017|incl|excl"]["va_total"], 157531)
    attendu_A = {
        # secteur      part_va  part_credit  ifr    ecart
        "BTP":        (8.52,   21.53,      2.53,   5767),
        "Pêche":      (11.61,  18.81,      1.62,   3193),
        "Commerce":   (21.06,  23.73,      1.13,   1182),
        "Services":   (29.83,  22.03,      0.74,  -3458),
        "Industries": (28.97,  13.90,      0.48,  -6684),
    }
    for nom, (pv, pc, ifr, ec) in attendu_A.items():
        s = A[nom]
        check(f"A {nom} part VA", d2(s["part_va"] * 100), pv)
        check(f"A {nom} part credit", d2(s["part_credit"] * 100), pc)
        check(f"A {nom} IFR", d2(s["ifr"]), ifr)
        check(f"A {nom} ecart", d2(s["ecart"], 0), ec)

    # --- section 2 : scenario B (hors extraction) ------------------------------
    B = {s["nom"]: s for s in scenarios["2017|excl|excl"]["secteurs"]}
    attendu_B = {
        "BTP": (10.13, 2.12), "Pêche": (13.81, 1.36), "Commerce": (25.04, 0.95),
        "Industries (manuf.)": (15.54, 0.89), "Services": (35.47, 0.62),
    }
    for nom, (pv, ifr) in attendu_B.items():
        check(f"B {nom} part VA", d2(B[nom]["part_va"] * 100), pv)
        check(f"B {nom} IFR", d2(B[nom]["ifr"]), ifr)
    check_eq("B : le sous-financement passe aux Services",
             scenarios["2017|excl|excl"]["secteurs"][-1]["nom"], "Services")
    check_eq("A : le sous-financement porte sur les Industries",
             scenarios["2017|incl|excl"]["secteurs"][-1]["nom"], "Industries")

    # --- invariance de l'ecart quand le residu est reparti ---------------------
    # L'ecran 2 affirme que l'ecart en MRU ne depend pas de l'hypothese sur le
    # residu. Demonstration : ecart_i = credit_i - part_va_i x base ; en repartissant
    # le residu au poids economique, credit_i gagne residu x part_va_i et la base
    # gagne residu, donc les deux termes ajoutes s'annulent. Verifie ici sur les
    # six combinaisons annee x extraction, pas seulement sur 2017.
    for annee in ANNEES_REF:
        for extraction in ("incl", "excl"):
            ex = {s["id"]: s for s in scenarios[f"{annee}|{extraction}|excl"]["secteurs"]}
            rp = {s["id"]: s for s in scenarios[f"{annee}|{extraction}|repart"]["secteurs"]}
            for sid in ex:
                check(f"ecart invariant au residu [{annee}|{extraction}|{sid}]",
                      rp[sid]["ecart"], ex[sid]["ecart"], tol=1e-9)
                check_eq(f"etiquette invariante au residu [{annee}|{extraction}|{sid}]",
                         rp[sid]["ecart_aff"], ex[sid]["ecart_aff"])

    # --- structure temporelle --------------------------------------------------
    dec20 = next(m for m in bloc_a if m["iso"] == "2020-12")
    part_ct = dec20["ct"] / (dec20["ct"] + dec20["mlt"]) * 100
    check("part court terme dec. 2020 (%)", d2(part_ct, 1), 63.7)
    check("encours total dec. 2020 = TOTAL bloc B", dec20["ct"] + dec20["mlt"], total_credit, tol=0.01)
    check("CT dec. 2020 = TOTAL bloc B", dec20["ct"], ct_total, tol=0.01)
    check("MLT dec. 2020 = TOTAL bloc B", dec20["mlt"], mlt_total, tol=0.01)
    p, d = bloc_a[0], bloc_a[-1]
    var_ct = (d["ct"] / p["ct"] - 1) * 100
    var_mlt = (d["mlt"] / p["mlt"] - 1) * 100
    check("variation CT (%)", d2(var_ct, 1), -4.2)
    check("variation MLT (%)", d2(var_mlt, 1), 16.3)

    # --- angles morts ----------------------------------------------------------
    agro = sum(float(va2017[b]) for b in ("Agriculture", "Elevage et chasse", "Sylviculture"))
    check("VA agro-pastorale 2017", agro, 35513)
    check("part agro-pastorale du PIB cf (%)", d2(agro / pib_cf * 100, 1), 15.5)
    couvert = scenarios["2017|incl|excl"]["va_total"]
    check("couvert + non couvert = PIB cf",
          couvert + sum(float(va2017[b]) for b in non_couvert), pib_cf, tol=0.5)

    # --- FINAN -----------------------------------------------------------------
    f04 = next(r for r in finan if r["annee"] == 2004)
    f19 = next(r for r in finan if r["annee"] == 2019)
    check("agences bancaires 2004", f04["banques_agences"], 23)
    check("agences bancaires 2019", f19["banques_agences"], 291)
    check("agences IMF 2004", f04["imf_agences"], 2)
    check("agences IMF 2019", f19["imf_agences"], 16)
    check("banques 2004", f04["banques_inst"], 8)
    check("banques 2019", f19["banques_inst"], 18)
    check("cooperatives 2013", next(r for r in finan if r["annee"] == 2013)["coop_inst"], 98)
    check("cooperatives 2014", next(r for r in finan if r["annee"] == 2014)["coop_inst"], 23)

    # --- ecriture --------------------------------------------------------------
    branches_non_couvertes = sorted(
        ({"nom": b, "va": float(va2017[b])} for b in non_couvert),
        key=lambda x: -x["va"])

    data = {
        "meta": {
            "genere": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            "annee_defaut": ANNEE_DEFAUT,
            "annees": ANNEES_REF,
            # Couverture REELLE de la source, distincte des annees de reference
            # proposees dans l'interface : sans elle on affiche « 2015-2017 »
            # comme si les comptes nationaux commencaient en 2015.
            "va_couverture": [min(comptes), max(comptes)],
            "unite": "MRU millions",
            "sources": {
                "va": "Comptes Nationaux.xlsx — VA par branche, 2005-2017",
                "credit": "Crédit bancaire.xlsx — encours CT/MLT 2020-2021 et ventilation sectorielle",
                "inclusion": "FINAN.xlsx — institutions et agences, 2004-2019",
            },
        },
        "credit": {
            "total": round(total_credit, 2),
            "ventile": round(ventile, 2),
            "residu": round(residu["encours"], 2),
            "residu_part": round(residu["encours"] / total_credit, 6),
            "postes": postes_credit,
        },
        "scenarios": scenarios,
        "maturite": {
            "mois": bloc_a,
            "part_ct_dec2020": round(part_ct, 4),
            "var_ct": round(var_ct, 4),
            "var_mlt": round(var_mlt, 4),
        },
        "inclusion": {
            "lignes": finan,
            "rupture": {"annee": 2014, "serie": "coop_inst", "avant": 98, "apres": 23,
                        "note": "rupture de serie sur le nombre de cooperatives, non lissee"},
        },
        "couverture": {
            "pib_cout_facteurs": pib_cf,
            "annee": 2017,
            "couvert": round(couvert, 2),
            "non_couvert": round(pib_cf - couvert, 2),
            "agro_pastoral": round(agro, 2),
            "branches_non_couvertes": branches_non_couvertes,
        },
        "mapping": [{"poste": p, "branches": b} for p, b in postes_map.items()],
        "controles": {"total": len(CHECKS), "reussis": sum(1 for c in CHECKS if c[0])},
    }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    # Copie embarquee : evite fetch(), donc le dashboard s'ouvre aussi en file://
    # sans serveur. C'est l'assurance-demo le jour du jury.
    with open(EMBED, "w", encoding="utf-8") as f:
        f.write("/* Genere par pipeline/build_data.py — ne pas editer a la main. */\n")
        f.write("window.DATA = ")
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    print(f"{len(CHECKS)} controles, tous verts.")
    print(f"data.json      ({os.path.getsize(OUT) / 1024:.1f} Ko)")
    print(f"js/data.js     ({os.path.getsize(EMBED) / 1024:.1f} Ko)  [copie embarquee, file:// ok]")


if __name__ == "__main__":
    main()
